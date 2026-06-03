import tkinter as ari
from tkinter import ttk
from tkinter import messagebox
import openpyxl as op

window = ari.Tk()
window.title("Salon Appointment System")
window.configure(bg="#dea3bd")
wb = op.Workbook()
ws = wb.active

def sabrina():
    if not cname_entry.get() or not month_entry.get() or not day_entry.get() or not time_entry.get():
        messagebox.showwarning("Input Error", "All fields are required")
        return

    # MONTH VALIDATION
    if not month_entry.get().isdigit():
        messagebox.showerror("Error", "Month must be a number")
        return

    month = int(month_entry.get())

    if month < 1 or month > 12:
        messagebox.showerror("Error", "Invalid month")
        return

    # DAY VALIDATION
    if not day_entry.get().isdigit():
        messagebox.showerror("Error", "Day must be a number")
        return

    day = int(day_entry.get())

    if day < 1 or day > 31:
        messagebox.showerror("Error", "Invalid day")
        return

    # TIME VALIDATION
    if ":" not in time_entry.get() or time_entry.get() < "00:00" or time_entry.get() > "23:59" or "pm" in time_entry.get().lower() or "am" in time_entry.get().lower():
        messagebox.showerror("Error", "Use 24-hour format like 13:30")
        return

    workbook = op.load_workbook("DeGuzman_Database.xlsx")
    sheet = workbook.active

    # DUPLICATE CHECK
    for row in sheet.iter_rows(min_row=2, values_only=True):

        if (
            str(row[2]) == month_entry.get() and
            str(row[3]) == day_entry.get() and
            str(row[4]) == time_entry.get()
        ):

            messagebox.showerror(
                "Unavailable",
                "Appointment slot already taken"
            )
            return

    # SAVE DATA HERE






    workbook = op.load_workbook("DeGuzman_Database.xlsx")
    sheet = workbook.active

    # Append new data
    sheet.append([
        sheet.max_row,
        cname_entry.get(),
        month_entry.get(),
        day_entry.get(),
        time_entry.get()
    ])

    workbook.save("DeGuzman_Database.xlsx")

    # Refresh table
    for row in table.get_children():
        table.delete(row)

    for row in sheet.iter_rows(min_row=2,values_only=True):
        table.insert("", "end", values=row)

    cname_entry.delete(0, ari.END)
    month_entry.delete(0, ari.END)
    day_entry.delete(0, ari.END)
    time_entry.delete(0, ari.END)
def rihanna(event):
    selected = table.focus()

    values = table.item(selected, "values")

    cname_entry.delete(0, ari.END)
    month_entry.delete(0, ari.END)
    day_entry.delete(0, ari.END)
    time_entry.delete(0, ari.END)

    cname_entry.insert(0, values[1])
    month_entry.insert(0, values[2])
    day_entry.insert(0, values[3])
    time_entry.insert(0, values[4])
def taylor():
    selected_item = table.selection()

    if not selected_item:
        messagebox.showwarning("No Selection", "Select a record first")
        return

    item = table.item(selected_item)

    order_id = item["values"][0]

    workbook = op.load_workbook("DeGuzman_Database.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):

        if str(row[0].value) == str(order_id):

            row[1].value = cname_entry.get()
            row[2].value = month_entry.get()
            row[3].value = day_entry.get()
            row[4].value = time_entry.get()

            break

    workbook.save("DeGuzman_Database.xlsx")

    # refresh table
    for data in table.get_children():
        table.delete(data)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", "end", values=row)

    messagebox.showinfo("Updated", "Appointment updated successfully")

    cname_entry.delete(0, ari.END)
    month_entry.delete(0, ari.END)
    day_entry.delete(0, ari.END)
    time_entry.delete(0, ari.END)
def cardi():
    selected_item = table.selection()

    if not selected_item:
        messagebox.showwarning("No Selection", "Select a record first")
        return

    item = table.item(selected_item)

    order_id = item["values"][0]

    workbook = op.load_workbook("DeGuzman_Database.xlsx")
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):

        if str(row[0].value) == str(order_id):
            sheet.delete_rows(row[0].row, 1)
            break

    workbook.save("DeGuzman_Database.xlsx")

    # refresh table
    for data in table.get_children():
        table.delete(data)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", "end", values=row)

    messagebox.showinfo("Deleted", "Appointment deleted successfully")

    cname_entry.delete(0, ari.END)
    month_entry.delete(0, ari.END)
    day_entry.delete(0, ari.END)
    time_entry.delete(0, ari.END)


# Form Title
title = ari.Label(window, text="Salon Appointment System", font=("Times New Roman", 14, "bold"), bg="#dea3bd")
title.grid(row=0, column=0, columnspan=6)

# Frame
genframe = ari.Frame(window, bg="#d96f9e", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=7, padx=10, pady=10)

# Customer Name Entry
cname_entry = ari.Entry(genframe, font=("Poppins", 12))
cname_entry.grid(row=2, column=1, columnspan=2, padx=10, pady=(10, 0))

cname_label = ari.Label(genframe, text="Customer Name", font=("Poppins", 10, "italic"), bg="#d96f9e")
cname_label.grid(row=3, column=1, columnspan=2)

# Date Entry
month_entry = ari.Entry(genframe, font=("Poppins", 12))
month_entry.grid(row=2, column=3, columnspan=2, padx=10, pady=(10, 0))

month = ari.Label(genframe, text="Month", font=("Poppins", 10, "italic"), bg="#d96f9e")
month.grid(row=3, column=3, columnspan=2)

day_entry = ari.Entry(genframe, font=("Poppins", 12))
day_entry.grid(row=4, column=1, columnspan = 2, padx=10, pady=(10, 0))

day = ari.Label(genframe, text="Day", font=("Poppins", 10, "italic"), bg="#d96f9e")
day.grid(row=5, column=1, columnspan=2)

# time Entry
time_entry = ari.Entry(genframe, font=("Poppins", 12))
time_entry.grid(row=4, column=3, columnspan=2, padx=10, pady=(10, 0))

time = ari.Label(genframe, text="Time", font=("Poppins", 10, "italic"), bg="#d96f9e")
time.grid(row=5, column=3, columnspan=2)

# Buttons
submit_btn = ari.Button(window, text="Submit",width=10, font=("times new roman", 12, "bold"), bg="#769174", fg="white", command = sabrina)
submit_btn.grid(row=6, column=1, columnspan=2, padx=10, pady=10)

update_btn = ari.Button(window, text="Update",width=10, font=("times new roman", 12, "bold"), bg="#516d9f", fg="white", command = taylor)
update_btn.grid(row=6, column=2, columnspan=2, padx=10, pady=10)

delete_btn = ari.Button(window, text="Delete",width=10, font=("times new roman", 12, "bold"), bg="#990000", fg="white", command = cardi)
delete_btn.grid(row=6, column=3, columnspan=2, padx=10, pady=10)

# Table
table = ttk.Treeview(
    window,
    columns=("Order ID", "Customer Name", "Month", "Day", "Time"),
    show="headings"
)

for headings in ("Order ID", "Customer Name", "Month", "Day", "Time"):
    table.heading(headings, text=headings)

table.grid(row=7, column=0, columnspan=6, padx=10, pady=10)
table.bind("<ButtonRelease-1>", rihanna)

# Excel headers
ws.append(["Order ID", "Customer Name", "Month", "Day", "Time"])
wb.save("DeGuzman_Database.xlsx")
window.mainloop()